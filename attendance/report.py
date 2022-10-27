import os
import openpyxl
from openpyxl.styles.borders import Border, Side
from attendance.model_turn import Turn

from user.Model import User
from .model import Register, Attendance
from peewee import fn
import os
from PyQt5.QtWidgets import QFileDialog
from PyQt5.QtCore import QDir

BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
CENTER = openpyxl.styles.Alignment(
    horizontal="center",
    vertical="center",
)

FONT_BOLD = openpyxl.styles.Font(bold=True)


class AttendanceReport:
    @staticmethod
    def open_save_dialog(file_name):
        filename = QFileDialog.getSaveFileName(
            None,
            "Guardar reporte",
            QDir.homePath() + "/" + file_name,
            "Excel (*.xlsx)",
        )
        return filename[0] if filename[0] != "" else None

    @staticmethod
    def generate(dt_start, dt_end, turn_id):
        try:
            turn = Turn.get_by_id(turn_id)
            file_name = f"Reporte_{dt_start.strftime('%d-%m-%Y')}__{dt_end.strftime('%d-%m-%Y')}_{turn.name}.xlsx"
            path = AttendanceReport.open_save_dialog(file_name)
            if not path:
                return

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Asistencias"

            ws.column_dimensions["A"].width = 5
            ws.column_dimensions["B"].width = 20
            row_start = 5
            col_start = 1

            # Header
            ws.merge_cells(
                start_row=row_start - 2,
                end_row=row_start - 2,
                start_column=col_start,
                end_column=col_start + 10,
            )
            ws.cell(row=row_start - 2, column=col_start).value = "Sistema de asistencia"
            ws.merge_cells(
                start_row=row_start - 1,
                end_row=row_start - 1,
                start_column=col_start,
                end_column=col_start + 10,
            )
            ws.cell(
                row=row_start - 1, column=col_start
            ).value = f'Reporte de asistencias del {dt_start.strftime("%d-%m-%Y")} al {dt_end.strftime("%d-%m-%Y")} del turno "{turn.name}"'

            users = (
                User.select()
                .where(User.turn == turn_id)
                .order_by(fn.Lower(User.lastname))
            )
            for i, user in enumerate(users):
                registers = user.registers.join(Attendance).where(
                    (Attendance.date >= dt_start) & (Attendance.date <= dt_end)
                )

                ws.cell(
                    row=row_start + i + 1,
                    column=col_start,
                    value=f"{i+1}",
                ).border = BORDER
                ws.cell(
                    row=row_start + i + 1,
                    column=col_start + 1,
                    value=f"{user.lastname}, {user.firstname}",
                ).border = BORDER

                for j, register in enumerate(registers):

                    if i == 0:
                        col = col_start + j + 2
                        val = register.attendance.date.day
                        cell = ws.cell(row=row_start, column=col)
                        cell.value = val
                        cell.border = BORDER
                        cell.alignment = CENTER
                        cell.font = FONT_BOLD
                        ws.column_dimensions[str(chr(64 + col))].width = 4

                    cell = ws.cell(row=row_start + i + 1, column=col_start + j + 2)
                    cell.value = (
                        "A"
                        if register.status == "Asistencia"
                        else ("T" if register.status == "Tardanza" else "I")
                    )
                    cell.border = BORDER
                    cell.alignment = CENTER

                c_attendance = registers.where(Register.status == "Asistencia").count()
                c_tardy = registers.where(Register.status == "Tardanza").count()
                c_absence = registers.where(Register.status == "Inasistencia").count()

                s_col = registers.count() + col_start + 2
                s_row = row_start + i + 1

                if i == 0:
                    a = ws.cell(row=row_start, column=s_col, value="Asistencias")
                    t = ws.cell(row=row_start, column=s_col + 1, value="Tardanzas")
                    i = ws.cell(row=row_start, column=s_col + 2, value="Inasistencias")
                    a.alignment = CENTER
                    t.alignment = CENTER
                    i.alignment = CENTER
                    a.border = BORDER
                    t.border = BORDER
                    i.border = BORDER
                    a.font = FONT_BOLD
                    t.font = FONT_BOLD
                    i.font = FONT_BOLD
                    ws.column_dimensions[str(chr(64 + s_col))].width = 10
                    ws.column_dimensions[str(chr(64 + s_col + 1))].width = 10
                    ws.column_dimensions[str(chr(64 + s_col + 2))].width = 12

                a = ws.cell(row=s_row, column=s_col, value=c_attendance)
                t = ws.cell(row=s_row, column=s_col + 1, value=c_tardy)
                i = ws.cell(row=s_row, column=s_col + 2, value=c_absence)

                a.alignment = CENTER
                t.alignment = CENTER
                i.alignment = CENTER
                a.border = BORDER
                t.border = BORDER
                i.border = BORDER

            wb.save(path)
            os.startfile(path)
        except Exception as e:
            print(e)
            # window.show_message("Error", str(e))
